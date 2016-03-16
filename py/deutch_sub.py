# encoding=utf8 

from openpyxl import load_workbook
import os
import re
import sys  

reload(sys)  
sys.setdefaultencoding('utf8')

# looking for language id, that is set in the 1-st row
def FindSpoken(lang, sheetName):
	for col in range(1,50):
		findSpokenCoord = sheetName.cell(row = 1, column = col)
		findSpokenVal = findSpokenCoord.value
		if findSpokenVal == lang:
			return col

def BatchLowcase(val):
	val = re.sub(ur'(?<!\w\.\w.)(?<![A-Z][a-z]\.)(?<=\.|\?|\!)\s', '\n', val)
	valSentenceList = val.split('\n')
	valSentenceList = filter(None, valSentenceList)
	newVal = ''
	for valSentence in valSentenceList:
		valWordsList = valSentence.split(' ')
		#print valWordsList
		valWordsList = filter(None, valWordsList)
		#print valWordsList
		
		#newVal += valWordsList[0] + ' '
		#del valWordsList[0]
		#print valWordsList
		for val in valWordsList:
			if not val[0].isupper():
				val = re.sub(ur'b(?=(\.|\,|\s|\Z))', 'p', val, flags=re.I)
				val = re.sub(ur'z', 'ts', val, flags=re.I)
				val = re.sub(ur'(?<=[aou])ch', 'kh', val, flags=re.I)
				val = re.sub(ur'(?<!t|s)s(?!(s|ch))', 'z', val, flags=re.I)
				val = re.sub(ur'sch', 'sh', val, flags=re.I)
				val = re.sub(ur'(?<=[eiyäö])ch', 'sch', val, flags=re.I)
				val = re.sub(ur'c(?=[eiyäö])', 'ts', val, flags=re.I)
				val = re.sub(ur'c(?=[auo])', 'k', val, flags=re.I)
				val = re.sub(ur'ig(?=(\.|\,|\s|\Z))', 'isch', val, flags=re.I)
				val = re.sub(ur'j', 'y', val, flags=re.I)
				val = re.sub(ur'qu', 'kv', val, flags=re.I)
				val = re.sub(ur'ss', 's', val, flags=re.I)
				val = re.sub(ur'th', 't', val, flags=re.I)
				val = re.sub(ur'v', 'f', val, flags=re.I)
				val = re.sub(ur'ow(?=(\.|\,|\s|\Z))', 'oh', val, flags=re.I)
				val = re.sub(ur'w', 'v', val, flags=re.I)
				val = re.sub(ur'ß', 's', val, flags=re.I)
				val = re.sub(ur'ä', 'ai', val, flags=re.I)
				val = re.sub(ur'ö', 'ue', val, flags=re.I)
				val = re.sub(ur'ü', 'ui', val, flags=re.I)
				#val = re.sub(ur'(?<!lor)d\b', '', val, flags=re.I)
				newVal += val + ' '
			else:
				newVal += val + ' '
	return newVal


# deleting tags if there are already present in deu or fra columns
# for applying corrections
def MatchSub(val):
	if val:
		val = BatchLowcase(val)
		return val

# tagging itself
def TagDoc(docPath):
	workBook = load_workbook(docPath)
	# sheets are not named correctly
	# so I'm iterating through all of them
	sheetList = workBook.get_sheet_names()
	val = ''
	for sheet in sheetList:
		# check for tags in excel file, if no, no save
		sheetName = workBook[sheet]
		
		# pattern for spoken text column naming
		intSpoken = 'SpokenText'
		fraSpoken = 'FRA.SpokenText'
		deuSpoken = 'DEU.SpokenText'

		intColNum = FindSpoken(intSpoken, sheetName)
		fraColNum = FindSpoken(fraSpoken, sheetName)
		deuColNum = FindSpoken(deuSpoken, sheetName)

		# sometimes int spoken text is spelled with a space. If yes, it will return None.
		# this is int spoken text spelling check
		if intColNum:
			# even for comments it is enough length
			for row in range(3,200):
				intRowCoord = sheetName.cell(row = row, column = intColNum)
				intRowVal = intRowCoord.value
				# if cell is not empty
				if intRowVal:
					fraRowCoord = sheetName.cell(row = row, column = fraColNum)	
					fraRowVal = fraRowCoord.value
					deuRowCoord = sheetName.cell(row = row, column = deuColNum)
					deuRowVal = deuRowCoord.value
					#print deuRowVal
					#intRowValList = ListFromVal(intRowVal)
					#fraRowValList = ListFromVal(fraRowVal)
					#deuRowValList = ListFromVal(deuRowVal)

					#intFraRowVal = MatchSub(intRowVal)
					#tagFraRowVal = MatchSub(fraRowVal)
					tagDeuRowVal = MatchSub(deuRowVal)
					#print tagDeuRowVal, '\n'
					# adding tagged content back to cells
					#intRowCoord.value = intFraRowVal
					#fraRowCoord.value = tagFraRowVal
					deuRowCoord.value = tagDeuRowVal
	workBook.save(docPath)
	print 'saved', docPath

# default path for docs on my PC for sh8 game
docDir =  "d:/svn/ue3/SH8Game/Production/Dialogs/"
#docDir =  "d:/sh8/xlsx_python_tests/"

# collecting all .xlsxs from supplied path
for path, dirs, fileNames in os.walk(docDir):
    for fileName in fileNames:
    	docPath = os.path.join(path, fileName)
    	if '.xlsx' in docPath:
    		TagDoc(docPath)
