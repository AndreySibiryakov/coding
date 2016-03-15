from openpyxl import load_workbook
import os

# list of text to search for
keyWordList = ['Resume', 'Label', 'Description', 'ClueText', 'Title', 'QTEtitle']

# default path for docs on my PC for sh8 game xlsx documents
#docDir =  "d:/svn/ue3/SH8Game/Production/Dialogs/"
docDir =  "d:/svn/ue3/SH8Game/Production/Data/"
#docDir =  "d:/sh8/xlsx_python_tests/"

# output for the log file
logFile = 'd:/sh8/xlsx_python_tests/genlog.txt'

# searching for INT column ID
# returns column serial nubmer
def FindBase(sheetName, keyWord):
	for col in range(1,50):
		findSpokenCoord = sheetName.cell(row = 1, column = col)
		findSpokenVal = findSpokenCoord.value
		if findSpokenVal == keyWord:
			return col

# searching for all localization columns that is present
# returns list of columns serial number 
def FindLoc(sheetName, keyWord):
	TextColList = []
	for col in range(1,100):
		findSpokenCoord = sheetName.cell(row = 1, column = col)
		findSpokenVal = findSpokenCoord.value
		#print findSpokenVal
		if findSpokenVal:
			if ('.' + keyWord) in findSpokenVal:
				TextColList.append(col)
	return TextColList

# comparing INT cell content with localization content
# returns string if INT and LOC cell are indentical
# returns string if LOC is empty while INT is not
def FindAndLog(docPath, keyWordList):
	# declaring var for storing log
	logVal = ''
	workBook = load_workbook(docPath)
	# for test purposes
	print docPath
	# obtaining list of all sheets in document
	sheetList = workBook.get_sheet_names()
	# adding path to log
	logVal += docPath + '\n'
	# iterating through key words
	for keyWord in keyWordList:
		# iterating through sheets in document
		for sheet in sheetList:
			sheetName = workBook[sheet]
			intColNum = FindBase(sheetName, keyWord)
			locColNumList = FindLoc(sheetName, keyWord)
			# checking if INT keyword is present in document
			if intColNum:
				# even for comments it is enough length
				for row in range(4,200):
					intRowCoord = sheetName.cell(row = row, column = intColNum)
					# obtaining INT cell value 
					intRowVal = intRowCoord.value				
					# checking if INT cell is not empty
					if intRowVal:
						# iterating through LOC columns in list
						for col in locColNumList:	
							locRowCoord = sheetName.cell(row = row, column = col)
							# obtaining LOC cell value 
							locRowVal = locRowCoord.value
							# checking whether LOC cell is duplicate of INT
							if intRowVal == locRowVal:
								#convering non ASCII characters
								#locASCII = str(intRowVal).encode('ascii', 'ignore').decode('ascii')
								#print intRowVal
								logVal += str(locRowCoord) + str(intRowVal) + '\n'
							# checking if LOC cell is empty while INT cell is not
							elif locRowVal == None:
								logVal += str(locRowCoord) + ' is empty\n'
	return logVal

# collecting all .xlsxs from supplied path
genLog = ''
for path, dirs, fileNames in os.walk(docDir):
    for fileName in fileNames:
    	docPath = os.path.join(path, fileName)
    	# filtering files except .xlsx
    	if '.xlsx' in docPath:
    		# filling log
    		genLog += FindAndLog(docPath, keyWordList)

# writing and saving the log file
filePath = open(logFile, 'wb')
filePath.write(genLog)
filePath.close()

