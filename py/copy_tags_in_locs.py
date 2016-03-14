from openpyxl import load_workbook
import os

# looking for language id, that is set in the 1-st row
def FindSpoken(lang, sheetName):
	for col in range(1,50):
		findSpokenCoord = sheetName.cell(row = 1, column = col)
		findSpokenVal = findSpokenCoord.value
		if findSpokenVal == lang:
			return col
		#else:
			#print 'No %s found' % lang
# creating list from cell content, words
def ListFromVal(val, int = True):
	#val.encode('utf-8')
	valList = val.split(' ')
	# filtering double spaces
	valList = filter(None, valList)
	# deleting existing tags in localization cells
	if int == False:
		valList = DeleteTagsFromVal(valList)
	return valList

# deleting tags if there are already present in deu or fra columns
# for applying corrections
def DeleteTagsFromVal(valList):
	newList = []
	for val in valList:
		if '/' not in val:
			newList.append(val)
	return newList

# creating dictionary from word serial number and word itself
def CreateTagDict(valList):
	wordCount = 0
	tagDict = {}
	for word in valList:
		if '/' in word:
			# constructing dictionary and substracting serial number of tags from total word count 
			tagDict[wordCount-(len(tagDict))] = word
		wordCount +=1
	return tagDict

def InsertTag(locValList, tagDict, multyplier):
	# correcting tag shift with multyplier if words quantity differs from int
	tagCount = 0
	for tagCountNum in tagDict:
		# applying offset for tags based on words quantity in loc cells
		tagMultiplied = int((tagCountNum*multyplier)+tagCount)			
		locValList.insert(tagMultiplied, tagDict[tagCountNum])
		tagCount += 1
	# creating string from list
	taggedVal = ' '.join(locValList)
	return taggedVal


def Multiplier(intRowVal, langRowVal, tagDict):
	# substracting tags from overall quantity of words in int cell
	# to get the initial word quantity
	noTagsIntLength = len(intRowVal)-len(tagDict)
	return (float(len(langRowVal))/float(noTagsIntLength))


# tagging itself
def TagDoc(docPath):
	workBook = load_workbook(docPath)
	# sheets are not named correctly
	# so I'm iterating through all of them
	sheetList = workBook.get_sheet_names()
	tagsInDocBool = False
	for sheet in sheetList:
		# check for tags in excel file, if no, no save
		sheetName = workBook[sheet]
		
		# pattern for spoken text column naming
		intSpoken = 'SpokenText'
		fraSpoken = 'FRA.SpokenText'
		deuSpoken = 'DEU.SpokenText'

		intColNum = FindSpoken(intSpoken, sheetName)
		fraColNum = FindSpoken(fraSpoken, sheetName)
		deuColNum = FindSpoken(deuSpoken, sheetName)

		# sometimes int spoken text is spelled with a space. If yes, it will return None.
		# this is int spoken text spelling check
		if intColNum:
			# even for comments it is enough length
			for row in range(1,200):
				intRowCoord = sheetName.cell(row = row, column = intColNum)
				intRowVal = intRowCoord.value
				# if cell is not empty
				if intRowVal:
					# all tags start from slash, so, this condition will skip all cells without tags
					if '/' in intRowVal:
						tagsInDocBool = True
						# getting fra and deu cell content
						fraRowCoord = sheetName.cell(row = row, column = fraColNum)	
						fraRowVal = fraRowCoord.value
						deuRowCoord = sheetName.cell(row = row, column = deuColNum)
						deuRowVal = deuRowCoord.value

						intRowValList = ListFromVal(intRowVal, int = True)
						fraRowValList = ListFromVal(fraRowVal, int = False)
						deuRowValList = ListFromVal(deuRowVal, int = False)

						intTagDict = CreateTagDict(intRowValList)
						
						fraRowValList = DeleteTagsFromVal(fraRowValList)
						deuRowValList = DeleteTagsFromVal(deuRowValList)

						fraMultiplier = Multiplier(intRowValList, fraRowValList, intTagDict)
						deuMultiplier = Multiplier(intRowValList, deuRowValList, intTagDict)

						tagFraRowVal = InsertTag(fraRowValList, intTagDict, fraMultiplier)
						tagDeuRowVal = InsertTag(deuRowValList, intTagDict, deuMultiplier)
						# adding tagged content back to cells
						fraRowCoord.value = tagFraRowVal
						deuRowCoord.value = tagDeuRowVal
				
		else:
			print 'not found INT SpokenText in', sheet, docPath
	if tagsInDocBool == True:
		workBook.save(docPath)
		print 'saved', docPath
	else:
		print 'no tags in', docPath

# default path for docs on my PC for sh8 game

docDir =  "d:/svn/ue3/SH8Game/Production/Dialogs/"
#docDir =  "d:/sh8/xlsx_python_tests/"

# collecting all .xlsxs from supplied path
for path, dirs, fileNames in os.walk(docDir):
    for fileName in fileNames:
    	docPath = os.path.join(path, fileName)
    	if '.xlsx' in docPath:
    		TagDoc(docPath)
