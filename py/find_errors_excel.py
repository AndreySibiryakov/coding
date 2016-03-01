# encoding=utf8 

from openpyxl import load_workbook
import os
import re
import sys  

reload(sys)  
sys.setdefaultencoding('utf8')

# tagging itself
def TagDoc(docPath):
	workBook = load_workbook(docPath)
	print docPath
	# sheets are not named correctly
	# so I'm iterating through all of them
	sheetList = workBook.get_sheet_names()
	val = ''
	for sheet in sheetList:
		# check for tags in excel file, if no, no save
		sheetName = workBook[sheet]
		# sometimes int spoken text is spelled with a space. If yes, it will return None.
		# this is int spoken text spelling check
		# even for comments it is enough length
		for row in range(1,400):
			for col in range(1,400):
				coord = sheetName.cell(row = row, column = col)
				val = coord.value
				if val:
					if '_FilterDatabase' in str(val):
						print 'is in ', str(coord)

# default path for docs on my PC for sh8 game
docDir =  "d:/svn/ue3/SH8Game/Production/Dialogs/"
#docDir =  "d:/sh8/xlsx_python_tests/"

# collecting all .xlsxs from supplied path
for path, dirs, fileNames in os.walk(docDir):
    for fileName in fileNames:
    	docPath = os.path.join(path, fileName)
    	if '.xlsx' in docPath:
    		TagDoc(docPath)
