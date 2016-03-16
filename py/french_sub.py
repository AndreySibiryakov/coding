# encoding=utf8 

from openpyxl import load_workbook
import os
import re
import sys  

reload(sys)  
sys.setdefaultencoding('utf8')

# looking for language id, that is set in the 1-st row
def FindSpoken(lang, sheetName):
	for col in range(1,50):
		findSpokenCoord = sheetName.cell(row = 1, column = col)
		findSpokenVal = findSpokenCoord.value
		if findSpokenVal == lang:
			return col

def BatchLowcase(val):
	val = re.sub(ur'(?<!\w\.\w.)(?<![A-Z][a-z]\.)(?<=\.|\?|\!)\s', '\n', val)
	valSentenceList = val.split('\n')
	valSentenceList = filter(None, valSentenceList)
	newVal = ''
	for valSentence in valSentenceList:
		valWordsList = valSentence.split(' ')
		#print valWordsList
		valWordsList = filter(None, valWordsList)
		#print valWordsList
		
		#newVal += valWordsList[0] + ' '
		#del valWordsList[0]
		#print valWordsList
		for val in valWordsList:
			if not val[0].isupper():
				val = re.sub(ur't\b', '', val, flags=re.I)
				val = re.sub(ur'(?<!\b)s\b(?! [aeyuio])', '', val, flags=re.I)
				val = re.sub(ur'z\b', '', val, flags=re.I)
				val = re.sub(ur'(?<!\b)d\b', '', val, flags=re.I)
				val = re.sub(ur'(?<![s])h(?!olmes)', '', val, flags=re.I)
				val = re.sub(ur'(?<=[eyioae])x\b', '', val, flags=re.I)
				val = re.sub(ur'p\b', '', val, flags=re.I)
				val = re.sub(ur'oi\b', 'va', val, flags=re.I)
				val = re.sub(ur'ois\b(?! [eyioae])', 'va', val, flags=re.I)
				val = re.sub(ur'x\b', '', val, flags=re.I)
				#val = re.sub(ur'(?<!lor)d\b', '', val, flags=re.I)
				newVal += val + ' '
			else:
				newVal += val + ' '
	return newVal


# deleting tags if there are already present in deu or fra columns
# for applying corrections
def MatchSub(val):
	if val:
		val = re.sub(ur'î', 'i', val, flags=re.I)
		val = re.sub(ur'ï(?=[àâäôéèëêïîùûüÿaeyuio])', 'j', val, flags=re.I)
		val = re.sub(ur'ç', 's', val, flags=re.I)
		val = re.sub(ur'ô', 'o', val, flags=re.I)
		val = re.sub(ur'û', 'u', val, flags=re.I)
		val = re.sub(ur'[èêëé]', 'e', val, flags=re.I)
		val = re.sub(ur'[àâ]', 'a', val, flags=re.I)
		val = re.sub(ur'\boui\b', 'vi', val, flags=re.I)
		val = re.sub(ur'est(?! [aeyuio])', 'e', val, flags=re.I)
		val = re.sub(ur'vous\b(?! [aeyuio])', 'voo', val, flags=re.I)
		val = re.sub(ur'uis\b(?! [aeyuio])', 'vi', val, flags=re.I)
		val = re.sub(ur'ui\b(?! [aeyuio])', 'vi', val, flags=re.I)
		val = re.sub(ur'es(?! [aeyuio])', 'e', val, flags=re.I)
		# was deleting too much
		#val = re.sub(ur'\w(?<![crlf0-9_])\b', '', val, flags=re.I)
		val = re.sub(ur'c(?=(\')|[ei])', 's', val, flags=re.I)
		val = re.sub(ur'c(?![ei])', 'k', val, flags=re.I)
		val = re.sub(ur'ch', 'sh', val, flags=re.I)
		val = re.sub(ur'g(?=[eiy])', 'j', val, flags=re.I)
		val = re.sub(ur'(?<!c)(qu|q)', 'k', val, flags=re.I)
		# returns s to the end of words
		val = re.sub(ur'Holme\b', 'Holmes', val, flags=re.I)
		val = re.sub(ur'(?<=i)lh', 'j', val, flags=re.I)
		val = re.sub(ur'(?<!i)lh', 'l', val, flags=re.I)
		val = re.sub(ur'sc(?=[eiy])', 's', val, flags=re.I)
		val = re.sub(ur'(ques|que)', 'k', val, flags=re.I)
		val = re.sub(ur'im(?![àâäôéèëêïîùûüÿaeyuiom])', 'em', val, flags=re.I)
		val = re.sub(ur'(?<=[àâäôéèëêïîùûüÿaeyuio])il', 'j', val, flags=re.I)
		val = re.sub(ur'(\b|(?<![àâäôéèëêïîùûüÿaeyuio]))(ei|eî)', 'e', val, flags=re.I)
		val = re.sub(ur'ail\b', 'aj', val, flags=re.I)
		val = re.sub(ur'aill\b', 'aj', val, flags=re.I)
		val = re.sub(ur'aille\b', 'aj', val, flags=re.I)
		val = re.sub(ur'eil\b', 'ey ', val, flags=re.I)
		val = re.sub(ur'eille\b', 'ey ', val, flags=re.I)
		val = re.sub(ur'cqu\b', 'ck ', val, flags=re.I)
		val = re.sub(ur'\bdes\b', 'de', val, flags=re.I)
		val = re.sub(ur'eau', 'o', val, flags=re.I)
		val = re.sub(ur'euil\b', 'ey ', val, flags=re.I)
		val = re.sub(ur'gn\b', 'n ', val, flags=re.I)
		val = re.sub(ur'(?<![eyioae])ier\b', 'e', val, flags=re.I)
		val = re.sub(ur'(?<![eyioae])ieu(\s|\.|\Z)', 'e', val, flags=re.I)
		val = re.sub(ur'(?<![eyioae])ill(?=[eyioae])', 'y', val, flags=re.I)
		val = re.sub(ur'(?<!\b)(?<!\')il\b', 'iy ', val, flags=re.I)
		val = re.sub(ur'(au|eau|ou)ld\b', 'o ', val, flags=re.I)
		val = re.sub(ur'(au|eau|ou)lt\b', 'o ', val, flags=re.I)
		val = re.sub(ur'\bm\.', 'monsieur', val, flags=re.I)
		val = re.sub(ur'\bune\b', 'un', val, flags=re.I)
		val = BatchLowcase(val)
		return val

# tagging itself
def TagDoc(docPath):
	workBook = load_workbook(docPath)
	# sheets are not named correctly
	# so I'm iterating through all of them
	sheetList = workBook.get_sheet_names()
	val = ''
	for sheet in sheetList:
		# check for tags in excel file, if no, no save
		sheetName = workBook[sheet]
		
		# pattern for spoken text column naming
		intSpoken = 'SpokenText'
		fraSpoken = 'FRA.SpokenText'
		deuSpoken = 'DEU.SpokenText'

		intColNum = FindSpoken(intSpoken, sheetName)
		fraColNum = FindSpoken(fraSpoken, sheetName)
		deuColNum = FindSpoken(deuSpoken, sheetName)

		# sometimes int spoken text is spelled with a space. If yes, it will return None.
		# this is int spoken text spelling check
		if intColNum:
			# even for comments it is enough length
			for row in range(3,200):
				intRowCoord = sheetName.cell(row = row, column = intColNum)
				intRowVal = intRowCoord.value
				# if cell is not empty
				if intRowVal:
					fraRowCoord = sheetName.cell(row = row, column = fraColNum)	
					fraRowVal = fraRowCoord.value
					deuRowCoord = sheetName.cell(row = row, column = deuColNum)
					deuRowVal = deuRowCoord.value
					#print deuRowVal
					#intRowValList = ListFromVal(intRowVal)
					#fraRowValList = ListFromVal(fraRowVal)
					#deuRowValList = ListFromVal(deuRowVal)

					#intFraRowVal = MatchSub(intRowVal)
					tagFraRowVal = MatchSub(fraRowVal)
					#tagDeuRowVal = MatchSub(deuRowVal)
					#print tagDeuRowVal, '\n'
					# adding tagged content back to cells
					#intRowCoord.value = intFraRowVal
					fraRowCoord.value = tagFraRowVal
					#deuRowCoord.value = tagDeuRowVal
	workBook.save(docPath)
	print 'saved', docPath

# default path for docs on my PC for sh8 game
docDir =  "d:/svn/ue3/SH8Game/Production/Dialogs/"
#docDir =  "d:/sh8/xlsx_python_tests/"

# collecting all .xlsxs from supplied path
for path, dirs, fileNames in os.walk(docDir):
    for fileName in fileNames:
    	docPath = os.path.join(path, fileName)
    	if '.xlsx' in docPath:
    		TagDoc(docPath)
